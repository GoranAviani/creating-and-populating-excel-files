import openpyxl
from openpyxl.styles import Font

test_title1 = ['keys', 'values']
test_data1 = {'key1': 'value1', 'key2': 'value2', 'key3': 'value3', 'key4': 'value4'}
test_data2 = {'column_name1': 'value1', 'column_name2': 'value2', 'column_name3': 'value3', 'column_name4': 'value4'}



def create_new_file():
    # Call a Workbook() function of openpyxl
    # to create a new blank Workbook object
    workbook = openpyxl.Workbook()

    # Get workbook active sheet
    # from the active attribute.
    sheet = workbook.active

    sheet_title = "business basic info"
    sheet.title = sheet_title

    print("active sheet title: " + sheet_title)

    # Note: The first row or column integer
    # is 1, not 0. Cell object is created by
    # using sheet object's cell() method.
    cell1 = sheet.cell(row=1, column=1)

    # writing values to cells
    cell1.value = "First name"

    c2 = sheet.cell(row=1, column=2)
    c2.value = "Last name"

    # Once have a Worksheet object, one can
    # access a cell object by its name also.
    # A2 means column = 1 & row = 2.
    c3 = sheet['A2']
    c3.value = "Goran"

    # B2 means column = 2 & row = 2.
    c4 = sheet['B2']
    c4.value = "Aviani"



    #create new sheet
    workbook.create_sheet(index=1, title="This_is_sheet2")
    sheet = workbook['This_is_sheet2']
    cell1 = sheet.cell(row=1, column=1)

    # writing values to cells
    cell1.value = "First name"

    c2 = sheet.cell(row=1, column=2)
    c2.value = "Last name"

    # create new sheet
    workbook.create_sheet(index=2, title="test_data1")
    sheet = workbook['test_data1']

    #writing list to rows
    for x in range(1, len(test_title1)+1):
        current_cell = sheet['A' + str(x)]
        current_cell.value = test_title1[x-1]


    #Creating simple table from simple dict
    workbook.create_sheet(index=3, title="test_data2")
    sheet = workbook['test_data2']
    test_data1_row = 5
    title_cell = sheet['C' + str(test_data1_row-1)]
    title_cell.value = 'This is test1 table'
    for key, value in test_data1.items():
        key_cell = sheet['C' + str(test_data1_row)]
        value_cell = sheet['D' + str(test_data1_row)]

        key_cell.value = key
        value_cell.value = value

        test_data1_row += 1



    #creating a simple table created from a dict with name as key
    test_data3 = {
        'Sweden_Pay_Now_Direct_debit': {'column_name1': 'value1', 'column_name2': 'value2', 'column_name3': 'value3',
                                        'column_name4': 'value4'},
        'Sweden_Pay_Now_Card': {'column_name1': 'value1', 'column_name2': 'value2', 'column_name3': 'value3',
                                        'column_name4': 'value4'}
    }
    workbook.create_sheet(index=4, title="test_data3")
    sheet = workbook['test_data3']
    row_counter = 15

    for key_title, value_table in test_data3.items():
        title_cell = sheet['B' + str(row_counter - 1)]
        title_cell.value = key_title
        title_cell.font = Font(bold=True)

        for key_column_name, value_result in test_data3[key_title].items():
            key_cell = sheet['B' + str(row_counter)]
            value_cell = sheet['C' + str(row_counter)]

            key_cell.value = key_column_name
            value_cell.value = value_result

            row_counter += 1

        row_counter += 2

    #ceating a vertical table with keys as columns and values as data in rows

    test_data4 = {
        '2020': {'column_name1': 'value1', 'column_name2': 'value2', 'column_name3': 'value3',
                                        'column_name4': 'value4'},
        '2021': {'column_name1': 'value1', 'column_name2': 'value2', 'column_name3': 'value3',
                                'column_name4': 'value4'}
    }

    workbook.create_sheet(index=5, title="test_data4")
    sheet = workbook['test_data4']
    row_counter = 4
    ALPHABET = ['A','B','C','D','F','G']
    alphabet_counter = 0
    for key_title, value_table in test_data4.items():
        title_cell = sheet['B' + str(row_counter)]
        title_cell.value = key_title
        row_counter += 1
        alphabet_counter = 0
        for key_column_name, value_result in test_data4[key_title].items():
            alphabet_counter += 1
            column_cell = sheet[ALPHABET[alphabet_counter] + '3']
            column_cell.value = key_column_name



    # Anytime you modify the Workbook object
    # or its sheets and cells, the spreadsheet
    # file will not be saved until you call
    # the save() workbook method.
    workbook.save("fist_example.xlsx")
